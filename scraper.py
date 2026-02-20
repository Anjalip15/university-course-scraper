"""
University & Course Data Scraper
=================================
Produces: university_course_data.xlsx
  - Sheet 1: Universities  (5 rows)
  - Sheet 2: Courses       (25 rows — 5 per university, mixed levels)

Scraping strategy:
  - Static seed data guarantees all 25 rows always exist
  - Live web scraping enriches duration/fees where possible
  - Country + level fallback used when pages are blocked
"""

import time
import requests
from bs4 import BeautifulSoup
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# SEED DATA — 5 universities × 5 courses (2 Bachelor's + 2 Master's + 1 PhD)
# ==============================================================================
UNIVERSITY_DATA = [
    {
        "university_name": "University of Texas at Austin",
        "country": "United States",
        "city": "Austin",
        "website": "https://www.utexas.edu",
        "courses": [
            {"course_name": "BS Computer Science",        "level": "Bachelor's", "discipline": "Computer Science",       "url": "https://catalog.utexas.edu/undergraduate/natural-sciences/degrees-and-programs/bs-computer-science/"},
            {"course_name": "BS Mathematics",             "level": "Bachelor's", "discipline": "Mathematics",             "url": "https://catalog.utexas.edu/undergraduate/natural-sciences/degrees-and-programs/bs-mathematics/"},
            {"course_name": "MS Computer Science",        "level": "Master's",   "discipline": "Computer Science",       "url": "https://www.cs.utexas.edu/graduate/prospective-students/masters-program"},
            {"course_name": "MS Data Science",            "level": "Master's",   "discipline": "Data Science",           "url": "https://ms-datascience.utexas.edu/"},
            {"course_name": "PhD Electrical Engineering", "level": "PhD",        "discipline": "Electrical Engineering",  "url": "https://www.ece.utexas.edu/academics/graduate/phd"},
        ]
    },
    {
        "university_name": "University of Toronto",
        "country": "Canada",
        "city": "Toronto",
        "website": "https://www.utoronto.ca",
        "courses": [
            {"course_name": "BSc Computer Science",        "level": "Bachelor's", "discipline": "Computer Science",      "url": "https://future.utoronto.ca/undergraduate-programs/computer-science/"},
            {"course_name": "BSc Life Sciences",           "level": "Bachelor's", "discipline": "Life Sciences",         "url": "https://future.utoronto.ca/undergraduate-programs/life-sciences/"},
            {"course_name": "MSc Artificial Intelligence", "level": "Master's",   "discipline": "Artificial Intelligence","url": "https://web.cs.toronto.edu/graduate/artificial-intelligence"},
            {"course_name": "MSc Statistics",              "level": "Master's",   "discipline": "Statistics",            "url": "https://www.statistics.utoronto.ca/graduate"},
            {"course_name": "PhD Immunology",              "level": "PhD",        "discipline": "Immunology",            "url": "https://immunology.utoronto.ca/graduate-studies/phd-program"},
        ]
    },
    {
        "university_name": "University of California, Berkeley",
        "country": "United States",
        "city": "Berkeley",
        "website": "https://www.berkeley.edu",
        "courses": [
            {"course_name": "BS Electrical Engineering & CS", "level": "Bachelor's", "discipline": "Electrical Engineering","url": "https://eecs.berkeley.edu/academics/undergraduate/eecs-bs/"},
            {"course_name": "BS Data Science",                "level": "Bachelor's", "discipline": "Data Science",          "url": "https://data.berkeley.edu/degrees/data-science-ba"},
            {"course_name": "MS Information & Data Science",  "level": "Master's",   "discipline": "Data Science",          "url": "https://ischool.berkeley.edu/programs/mids"},
            {"course_name": "MS Civil & Environmental Eng.",  "level": "Master's",   "discipline": "Civil Engineering",     "url": "https://ce.berkeley.edu/programs/grad"},
            {"course_name": "PhD Neuroscience",               "level": "PhD",        "discipline": "Neuroscience",          "url": "https://neuroscience.berkeley.edu/phd-program/"},
        ]
    },
    {
        "university_name": "University of Michigan",
        "country": "United States",
        "city": "Ann Arbor",
        "website": "https://umich.edu",
        "courses": [
            {"course_name": "BS Computer Science",        "level": "Bachelor's", "discipline": "Computer Science",       "url": "https://lsa.umich.edu/lsa/academics/majors-minors/computer-science.html"},
            {"course_name": "BS Mechanical Engineering",  "level": "Bachelor's", "discipline": "Mechanical Engineering", "url": "https://me.engin.umich.edu/academics/undergraduate/"},
            {"course_name": "MS Robotics",                "level": "Master's",   "discipline": "Robotics",              "url": "https://robotics.umich.edu/academics/courses/graduate-courses/"},
            {"course_name": "MS Applied Statistics",      "level": "Master's",   "discipline": "Statistics",            "url": "https://lsa.umich.edu/stats/graduate-students/graduate-programs/mas.html"},
            {"course_name": "PhD Biomedical Engineering", "level": "PhD",        "discipline": "Biomedical Engineering", "url": "https://bme.umich.edu/academics/graduate/phd/"},
        ]
    },
    {
        "university_name": "University of Edinburgh",
        "country": "United Kingdom",
        "city": "Edinburgh",
        "website": "https://www.ed.ac.uk",
        "courses": [
            {"course_name": "BSc Computer Science",        "level": "Bachelor's", "discipline": "Computer Science",      "url": "https://www.ed.ac.uk/studying/undergraduate/degrees/index.php?action=view&code=G400"},
            {"course_name": "BSc Mathematics",             "level": "Bachelor's", "discipline": "Mathematics",           "url": "https://www.ed.ac.uk/studying/undergraduate/degrees/index.php?action=view&code=G100"},
            {"course_name": "MSc Artificial Intelligence", "level": "Master's",   "discipline": "Artificial Intelligence","url": "https://www.ed.ac.uk/studying/postgraduate/degrees/index.php?r=site/view&id=107"},
            {"course_name": "MSc Data Science",            "level": "Master's",   "discipline": "Data Science",          "url": "https://www.ed.ac.uk/studying/postgraduate/degrees/index.php?r=site/view&id=902"},
            {"course_name": "PhD Informatics",             "level": "PhD",        "discipline": "Informatics",           "url": "https://www.ed.ac.uk/informatics/postgraduate/phd"},
        ]
    },
]

# Duration fallback by (level, country)
KNOWN_DURATIONS = {
    ("Bachelor's", "United States"):  "4 years",
    ("Bachelor's", "Canada"):         "4 years",
    ("Bachelor's", "United Kingdom"): "3 years",
    ("Master's",   "United States"):  "1-2 years",
    ("Master's",   "Canada"):         "1-2 years",
    ("Master's",   "United Kingdom"): "1 year",
    ("PhD",        "United States"):  "4-6 years",
    ("PhD",        "Canada"):         "4-5 years",
    ("PhD",        "United Kingdom"): "3-4 years",
}


# ==============================================================================
# SCRAPING HELPERS
# ==============================================================================
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def extract_duration(text, level, country):
    t = text.lower()
    if any(x in t for x in ["four year", "4 year", "four-year", "4-year"]):
        return "4 years"
    if any(x in t for x in ["three year", "3 year", "three-year", "3-year"]):
        return "3 years"
    if any(x in t for x in ["two year", "2 year"]):
        return "2 years"
    if any(x in t for x in ["one year", "1 year"]):
        return "1 year"
    return KNOWN_DURATIONS.get((level, country), "Refer official website")


def extract_fees(text):
    if "$" in text or "£" in text or "tuition" in text.lower():
        return "Refer official website (fees listed on page)"
    return "Refer official website"


def scrape_page(url, level, country):
    """Returns (duration, fees). Falls back gracefully on any failure."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            raise Exception(f"HTTP {r.status_code}")
        text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)
        return extract_duration(text, level, country), extract_fees(text)
    except Exception:
        return KNOWN_DURATIONS.get((level, country), "Refer official website"), "Refer official website"


# ==============================================================================
# BUILD DATA
# ==============================================================================
uni_rows, course_rows = [], []
course_num = 1

for u_idx, uni in enumerate(UNIVERSITY_DATA, 1):
    uid = f"U{u_idx:03}"
    uni_rows.append({
        "university_id":   uid,
        "university_name": uni["university_name"],
        "country":         uni["country"],
        "city":            uni["city"],
        "website":         uni["website"],
    })
    print(f"\n[{uid}] {uni['university_name']}")

    for course in uni["courses"]:
        cid = f"C{course_num:03}"
        duration, fees = scrape_page(course["url"], course["level"], uni["country"])

        course_rows.append({
            "course_id":    cid,
            "university_id": uid,
            "course_name":  course["course_name"],
            "level":        course["level"],
            "discipline":   course["discipline"],
            "duration":     duration,
            "fees":         fees,
            "eligibility":  "Refer official website",
        })
        print(f"  {cid} | {course['course_name']:<40} | {duration}")
        course_num += 1
        time.sleep(0.8)

assert len(course_rows) == 25, f"Expected 25 courses, got {len(course_rows)}"
print(f"\n{'='*50}")
print(f"  Universities : {len(uni_rows)}")
print(f"  Courses      : {len(course_rows)}")
print(f"{'='*50}\n")


# ==============================================================================
# EXCEL STYLES
# ==============================================================================
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "E8F4FD"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"
GREEN_STAT = "E2EFDA"

thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color=MID_BLUE)

LEVEL_COLORS = {
    "Bachelor's": "D6E4F0",
    "Master's":   "E2EFDA",
    "PhD":        "FCE4D6",
}


def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(bold=False):
    return Font(name="Arial", size=10, bold=bold)

def stat_val_font():
    return Font(name="Arial", size=11, bold=True, color=MID_BLUE)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, text, col_span):
    ws.merge_cells(f"A1:{get_column_letter(col_span)}1")
    cell = ws["A1"]
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def write_header_row(ws, row_num, headers):
    fill = PatternFill("solid", fgColor=MID_BLUE)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=c)
        cell.value     = h
        cell.font      = hdr_font()
        cell.fill      = fill
        cell.alignment = center()
        cell.border    = thick_border()
    ws.row_dimensions[row_num].height = 22


def write_data_row(ws, row_num, values, center_cols=(), alt=False):
    fill = PatternFill("solid", fgColor=ACCENT if alt else WHITE)
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c)
        cell.value     = v
        cell.font      = body_font()
        cell.fill      = fill
        cell.border    = thin_border()
        cell.alignment = center() if c in center_cols else left()
    ws.row_dimensions[row_num].height = 20


def write_summary_block(ws, start_row, col, stats):
    """Write a labelled summary panel. stats = list of (label, value)."""
    lc = get_column_letter(col)
    vc = get_column_letter(col + 1)

    # Header
    ws.merge_cells(f"{lc}{start_row}:{vc}{start_row}")
    hdr = ws.cell(row=start_row, column=col)
    hdr.value     = "Summary"
    hdr.font      = hdr_font(size=10)
    hdr.fill      = PatternFill("solid", fgColor=MID_BLUE)
    hdr.alignment = center()
    hdr.border    = thick_border()

    for r_offset, (label, value) in enumerate(stats, 1):
        label_cell = ws.cell(row=start_row + r_offset, column=col)
        val_cell   = ws.cell(row=start_row + r_offset, column=col + 1)

        label_cell.value     = label
        label_cell.font      = body_font(bold=True)
        label_cell.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
        label_cell.border    = thin_border()
        label_cell.alignment = left()

        val_cell.value     = value
        val_cell.font      = stat_val_font()
        val_cell.fill      = PatternFill("solid", fgColor=GREEN_STAT)
        val_cell.border    = thin_border()
        val_cell.alignment = center()

    ws.column_dimensions[lc].width = 22
    ws.column_dimensions[vc].width = 10


# ==============================================================================
# BUILD EXCEL
# ==============================================================================
wb = Workbook()

# ------------------------------------------------------------------------------
# SHEET 1 — Universities
# ------------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Universities"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

write_title(ws1, "🎓  University Dataset", col_span=5)

headers_u = ["University ID", "University Name", "Country", "City", "Official Website"]
write_header_row(ws1, 2, headers_u)

for i, row in enumerate(uni_rows, 3):
    write_data_row(ws1, i,
        [row["university_id"], row["university_name"],
         row["country"], row["city"], row["website"]],
        center_cols=(1,), alt=(i % 2 == 0)
    )
    # Clickable hyperlink
    link_cell          = ws1.cell(row=i, column=5)
    link_cell.hyperlink = row["website"]
    link_cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")

set_col_widths(ws1, [14, 36, 18, 16, 36])

# Count unique countries
countries = list({r["country"] for r in uni_rows})
write_summary_block(ws1, start_row=2, col=7, stats=[
    ("Total Universities", len(uni_rows)),
    ("Countries",          len(countries)),
])

# ------------------------------------------------------------------------------
# SHEET 2 — Courses
# ------------------------------------------------------------------------------
ws2 = wb.create_sheet("Courses")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

write_title(ws2, "📚  Course Dataset", col_span=8)

headers_c = ["Course ID", "University ID", "Course Name", "Level",
             "Discipline", "Duration", "Fees", "Eligibility"]
write_header_row(ws2, 2, headers_c)

for i, row in enumerate(course_rows, 3):
    write_data_row(ws2, i,
        [row["course_id"], row["university_id"], row["course_name"], row["level"],
         row["discipline"], row["duration"], row["fees"], row["eligibility"]],
        center_cols=(1, 2, 4, 6), alt=(i % 2 == 0)
    )
    # Colour-coded level badge
    level_cell       = ws2.cell(row=i, column=4)
    level_cell.fill  = PatternFill("solid", fgColor=LEVEL_COLORS.get(row["level"], WHITE))
    level_cell.font  = Font(name="Arial", size=10, bold=True)
    level_cell.alignment = center()

    # Yellow highlight for optional fields that fell back
    for col in (7, 8):
        cell = ws2.cell(row=i, column=col)
        if "Refer official website" in str(cell.value):
            cell.fill = PatternFill("solid", fgColor=YELLOW)

set_col_widths(ws2, [12, 14, 32, 14, 26, 14, 28, 24])

# Summary stats — computed from actual data
level_counts = Counter(r["level"] for r in course_rows)
write_summary_block(ws2, start_row=2, col=10, stats=[
    ("Total Courses",       len(course_rows)),
    ("Linked Universities", len(set(r["university_id"] for r in course_rows))),
    ("Bachelor's",          level_counts["Bachelor's"]),
    ("Master's",            level_counts["Master's"]),
    ("PhD",                 level_counts["PhD"]),
])

# ------------------------------------------------------------------------------
# SAVE
# ------------------------------------------------------------------------------
OUTPUT = "university_course_data.xlsx"
wb.save(OUTPUT)
print(f"✓ Saved: {OUTPUT}")
print(f"  Sheet 1 — Universities : {len(uni_rows)} rows")
print(f"  Sheet 2 — Courses      : {len(course_rows)} rows")
print(f"  Level breakdown        : {dict(level_counts)}")